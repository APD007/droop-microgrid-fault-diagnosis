"""Turn pilot_results.csv into a formatted Excel workbook."""

import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WHITE = Font(bold=True, color="FFFFFFFF", size=10)
THIN = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")

# column groups: (label, fill, [csv field names])
GROUPS = [
    ("SETUP", "FF404040",
     ["case", "desc", "R_A_a", "R_A_b", "R_A_c", "F1", "F2"]),
    ("DATASET COLUMNS - RMS", "FF2E75B6",
     ["V1_a", "V1_b", "V1_c", "I1_a", "I1_b", "I1_c",
      "V2_a", "V2_b", "V2_c", "I2_a", "I2_b", "I2_c"]),
    ("DC OFFSET - separates upper from lower", "FFC00000",
     ["I1mean_a", "I1mean_b", "I1mean_c", "I2mean_a", "I2mean_b", "I2mean_c"]),
    ("HALF-CYCLE RMS", "FFBF8F00",
     ["I1pos_a", "I1pos_b", "I1pos_c", "I1neg_a", "I1neg_b", "I1neg_c",
      "I2pos_a", "I2pos_b", "I2pos_c", "I2neg_a", "I2neg_b", "I2neg_c"]),
    ("DIAGNOSTICS", "FF548235",
     ["f1", "f2", "P1", "P2", "Q1", "Q2"]),
    ("RUN", "FF7030A0", ["wall_s", "ok"]),
]


def num(v):
    try:
        f = float(v)
        return round(f, 4)
    except (TypeError, ValueError):
        return v


def band(ws, row, c1, c2, label, rgb, size=10):
    fill = PatternFill("solid", fgColor=rgb)
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, Font(
            bold=True, color="FFFFFFFF", size=size), CTR, BOX
    if label is not None:
        ws.cell(row=row, column=c1).value = label
        if c2 > c1:
            ws.merge_cells(start_row=row, start_column=c1,
                           end_row=row, end_column=c2)


def sheet_results(wb, rows):
    ws = wb.create_sheet("Pilot Results")
    col = 1
    for label, rgb, fields in GROUPS:
        band(ws, 1, col, col + len(fields) - 1, label, rgb, size=10)
        for f in fields:
            c = ws.cell(row=2, column=col)
            c.value = f
            c.fill = PatternFill("solid", fgColor=rgb)
            c.font, c.alignment, c.border = WHITE, CTR, BOX
            col += 1

    for i, r in enumerate(rows, start=3):
        col = 1
        for _, _, fields in GROUPS:
            for f in fields:
                cell = ws.cell(row=i, column=col)
                cell.value = num(r.get(f, ""))
                cell.border = BOX
                if f not in ("desc", "F1", "F2"):
                    cell.alignment = CTR
                # highlight the faulted inverter's DC offset
                if f.startswith(("I1mean", "I2mean")) and isinstance(cell.value, float):
                    if abs(cell.value) > 1.0:
                        cell.fill = PatternFill("solid", fgColor="FFFFC7CE")
                        cell.font = Font(bold=True, color="FF9C0006")
                col += 1

    ws.freeze_panes = "C3"
    ws.row_dimensions[1].height = 22
    for c in range(1, col):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = 9
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    return ws


MAPPING = [
    # pulse, phase, half killed, device, mean I on that phase, pos/neg ratio
    (1, "b", "negative", "LOWER", 6.956, 15.75),
    (2, "b", "positive", "UPPER", -6.908, 0.07),
    (3, "c", "negative", "LOWER", 6.939, 14.92),
    (4, "c", "positive", "UPPER", -6.929, 0.06),
    (5, "a", "negative", "LOWER", 6.922, 15.29),
    (6, "a", "positive", "UPPER", -6.944, 0.07),
]


def sheet_mapping(wb):
    ws = wb.create_sheet("Pulse Mapping")
    ws["A1"] = "Pulse-to-IGBT mapping, measured (not assumed)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ("Determined by which phase current loses which half-cycle when that "
                "pulse is forced to zero. Identical on both inverters.")
    ws["A2"].font = Font(italic=True, size=10)

    hdr = ["Pulse", "Leg / phase", "Half-cycle lost", "Device",
           "mean I on that phase (A)", "pos/neg RMS ratio"]
    band(ws, 4, 1, len(hdr), None, "FF404040")
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=j).value = h

    for i, row in enumerate(MAPPING, start=5):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j)
            c.value = v
            c.border, c.alignment = BOX, CTR
        ws.cell(row=i, column=4).font = Font(bold=True)

    ws["A12"] = "Reasoning"
    ws["A12"].font = Font(bold=True, size=11)
    for i, t in enumerate([
        "Positive phase current is sourced by the inverter through the UPPER device.",
        "Negative phase current is sunk through the LOWER device.",
        "So a suppressed POSITIVE half-cycle (ratio << 1) means the UPPER switch is open,",
        "and a suppressed NEGATIVE half-cycle (ratio >> 1) means the LOWER switch is open.",
        "",
        "Note this is NOT the textbook 1=A-upper, 2=A-lower ordering. Legs come out in the",
        "order b, c, a and odd pulses are the lower devices. Measured, on both inverters.",
    ], start=13):
        ws.cell(row=i, column=1).value = t

    for col, w in zip("ABCDEF", [8, 12, 16, 10, 24, 18]):
        ws.column_dimensions[col].width = w
    return ws


def sheet_findings(wb):
    ws = wb.create_sheet("Findings")
    ws["A1"] = "Pilot findings - 15 runs, 0.4 s each, StopTime/Ts as delivered"
    ws["A1"].font = Font(bold=True, size=12)

    items = [
        ("Q1", "Does setVariable reach the model?", "PASS", "FF548235",
         "R=32 -> I1_a=4.309 | R=64 -> 2.521 | R=[16,32,32] -> 7.057/4.227/4.919. "
         "The load moves the currents and unbalance breaks phase symmetry, so the "
         "guarded InitFcn is working and the sweep will not produce identical rows."),
        ("Q2", "Is an open pulse visible at all?", "PASS", "FF548235",
         "Every fault produces a DC offset of 1.5-7 A on the faulted inverter against "
         "0.001-0.006 A when healthy. Unmistakable."),
        ("Q3", "Which pulse drives which IGBT?", "RESOLVED", "FF548235",
         "See the Pulse Mapping sheet. Consistent across both inverters."),
        ("Q4", "Are the 49 states separable on RMS alone?", "NO", "FFC00000",
         "Closest pair is DG1 pulse 3 vs DG1 pulse 4 - lower and upper of the SAME leg. "
         "Distance across the 12 RMS features is 0.0277, which is 4.9e-05 of the healthy "
         "feature magnitude. RMS discards sign, so a missing positive half-cycle and a "
         "missing negative half-cycle look almost identical. On the 28-column schema as "
         "written, 49 classes collapse to roughly 16 (healthy + 3 legs, per inverter)."),
        ("FIX", "Six extra columns solve it", "DECISION NEEDED", "FFBF8F00",
         "The mean (DC offset) of each inverter's three phase currents separates upper "
         "from lower cleanly: pulse 3 gives (-5.41, -1.53, +6.94) and pulse 4 gives "
         "(+5.40, +1.53, -6.93) - exact opposites. Distance 17.9 versus 0.0277 on RMS, "
         "about 650x more discriminative. Adding I1mean_a/b/c and I2mean_a/b/c takes the "
         "schema from 28 to 34 columns and needs your guide's approval."),
        ("SPEED", "Runtime is far better than projected", "GOOD", "FF2E75B6",
         "6.0 s per run, not the 33.6 s projected - because SaveOutput was switched off "
         "and logging decimated. 4704 runs is about 7.8 h on one process, or roughly 2 h "
         "split across four. Changing the load did not trigger an accelerator rebuild "
         "(case 14 took 6.0 s), so run ordering matters less than expected."),
    ]

    r = 3
    for tag, q, verdict, rgb, detail in items:
        ws.cell(row=r, column=1).value = tag
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        ws.cell(row=r, column=2).value = q
        ws.cell(row=r, column=2).font = Font(bold=True, size=11)
        c = ws.cell(row=r, column=3)
        c.value = verdict
        c.fill = PatternFill("solid", fgColor=rgb)
        c.font = Font(bold=True, color="FFFFFFFF", size=11)
        c.alignment = CTR
        ws.cell(row=r + 1, column=2).value = detail
        ws.cell(row=r + 1, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=3)
        ws.row_dimensions[r + 1].height = 62
        r += 3

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 95
    ws.column_dimensions["C"].width = 18
    return ws


def main():
    with open(ROOT / "pilot" / "pilot_results.csv", newline="") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()
    wb.remove(wb.active)
    sheet_findings(wb)
    sheet_results(wb, rows)
    sheet_mapping(wb)
    wb.save(ROOT / "pilot" / "pilot_results.xlsx")
    print(f"wrote pilot_results.xlsx  ({len(rows)} runs, {len(wb.sheetnames)} sheets)")
    for s in wb.sheetnames:
        print("   -", s)


if __name__ == "__main__":
    main()

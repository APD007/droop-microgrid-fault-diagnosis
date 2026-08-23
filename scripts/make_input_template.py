"""
Build the Excel template to hand over for testing.

The format question ("what columns do you need?") is best answered by giving
someone a sheet with the right headers already in it. This writes
docs/model_input_template.xlsx with four sheets:

    Instructions   what to fill in and how to produce it
    Measurements   the 42 headers, empty, ready to fill - one row per test case
    Example        three real rows so the format is unambiguous
    Waveforms      the alternative: paste raw time series instead, and let
                   features_from_waveforms.py do the reduction

Usage:
    python scripts/make_input_template.py
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import features as F

ROOT = Path(__file__).resolve().parent.parent

HDR = Font(bold=True, color="FFFFFFFF", size=10)
THIN = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")

GROUPS = [
    ("TRUE RMS", "FF2E75B6", F.MEASURED_RMS),
    ("DC OFFSET", "FFC00000", F.MEASURED_DC),
    ("FUNDAMENTAL MAGNITUDE", "FF548235",
     [c for c in F.MEASURED_FUND if "ang" not in c]),
    ("FUNDAMENTAL ANGLE", "FFBF8F00",
     [c for c in F.MEASURED_FUND if "ang" in c]),
]


def band(ws, row, c1, c2, label, rgb):
    fill = PatternFill("solid", fgColor=rgb)
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, HDR, CTR, BOX
    ws.cell(row=row, column=c1).value = label
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row,
                       end_column=c2)


def sheet_measurements(wb, rows=None, title="Measurements"):
    ws = wb.create_sheet(title)
    col = 1
    for label, rgb, cols in GROUPS:
        band(ws, 1, col, col + len(cols) - 1, label, rgb)
        for c in cols:
            cell = ws.cell(row=2, column=col)
            cell.value = c
            cell.fill = PatternFill("solid", fgColor=rgb)
            cell.font, cell.alignment, cell.border = HDR, CTR, BOX
            ws.column_dimensions[get_column_letter(col)].width = 11
            col += 1

    if rows is not None:
        flat = [c for _, _, cols in GROUPS for c in cols]
        for i, (_, r) in enumerate(rows.iterrows(), start=3):
            for j, c in enumerate(flat, start=1):
                cell = ws.cell(row=i, column=j)
                cell.value = round(float(r[c]), 5)
                cell.border, cell.alignment = BOX, CTR

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 20
    return ws


def sheet_instructions(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws["A1"] = "Model input template — droop microgrid fault diagnosis"
    ws["A1"].font = Font(bold=True, size=14)

    lines = [
        ("", ""),
        ("What this is for",
         "Filling this in lets you test the trained model on your own data. "
         "One row per operating condition."),
        ("", ""),
        ("Two ways to give us the data", ""),
        ("  Option A — summary values",
         "Fill the 'Measurements' sheet with the 42 numbers per test case. "
         "See 'Example' for three real rows."),
        ("  Option B — raw waveforms",
         "Easier. Paste sampled time series into the 'Waveforms' sheet and run "
         "features_from_waveforms.py, which computes the 42 numbers for you."),
        ("", ""),
        ("What the columns mean", ""),
        ("  V1_a … I2_c",
         "True RMS of the phase voltages and currents at both inverters, over "
         "a whole number of fundamental cycles."),
        ("  I1mean_a … I2mean_c",
         "The MEAN (DC offset) of each phase current over the same window. "
         "Near zero when healthy; several amps when a switch is open. This is "
         "what distinguishes an open upper switch from an open lower one."),
        ("  V1f_a … I2f_c",
         "Magnitude of the FUNDAMENTAL component only, RMS-scaled — not the "
         "true RMS, which includes harmonics."),
        ("  V1ang_a … I2ang_c",
         "Phase angle of that fundamental component, in degrees, all measured "
         "against V1_a. So V1ang_a is 0 by definition."),
        ("", ""),
        ("Important", ""),
        ("  Window",
         "Average over a whole number of fundamental cycles (4 is what the "
         "model was trained with). The frequency is NOT exactly 50 Hz — droop "
         "control puts it near 49.98 Hz, and it moves with load."),
        ("  Sampling",
         "If sending waveforms, sample at 20 kHz or faster. The carrier is "
         "10 kHz; below that the harmonic content the model reads is lost."),
        ("  Settling",
         "Take the window from a settled part of the run. Black start is at "
         "t = 0.04 s and the droop filter has a ~32 ms time constant, so "
         "anything after t = 0.2 s is safe."),
        ("  Units",
         "Volts, amps, seconds, degrees, hertz. Phase-to-ground voltages."),
        ("", ""),
        ("What you get back", ""),
        ("  Fault",
         "Which PWM pulse is open on each inverter, as twelve 1/0 flags, plus "
         "a confidence for each."),
        ("  Load",
         "R_a, R_b, R_c in ohms, and the degree of unbalance as a percentage."),
        ("", ""),
        ("If you also include the true answers",
         "Add columns R_a, R_b, R_c and DG1_PWM1..6, DG2_PWM1..6 and the "
         "script will score itself against them automatically."),
        ("", ""),
        ("To run it", "python scripts/predict.py <your file>.xlsx"),
    ]
    for i, (a, b) in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=a).font = Font(
            bold=not a.startswith("  ") and bool(a), size=10)
        c = ws.cell(row=i, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 88
    return ws


def sheet_waveforms(wb):
    ws = wb.create_sheet("Waveforms")
    cols = ["time"] + [f"{s}_{p}" for s in ("V1", "I1", "V2", "I2")
                       for p in F.PHASES] + ["f1"]
    band(ws, 1, 1, len(cols), "RAW WAVEFORMS — one row per SAMPLE, "
                              "one sheet per operating condition", "FF404040")
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=j)
        cell.value = c
        cell.fill = PatternFill("solid", fgColor="FF404040")
        cell.font, cell.alignment, cell.border = HDR, CTR, BOX
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.cell(row=4, column=1,
            value="time in seconds · instantaneous volts and amps · f1 optional "
                  "(estimated from V1_a if absent) · at least 4 cycles, "
                  "sampled at 20 kHz or faster").font = Font(italic=True, size=9)
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 20
    return ws


def main():
    src = ROOT / "data" / "dataset_full.csv"
    example = None
    if src.exists():
        d = pd.read_csv(src)
        # one healthy, one single fault, one double fault - the three shapes
        pw = [f"DG1_PWM{k}" for k in range(1, 7)] + \
             [f"DG2_PWM{k}" for k in range(1, 7)]
        n_open = 12 - d[pw].sum(axis=1)
        example = pd.concat([d[n_open == 0].head(1),
                             d[n_open == 1].head(1),
                             d[n_open == 2].head(1)])

    wb = Workbook()
    wb.remove(wb.active)
    sheet_instructions(wb)
    sheet_measurements(wb)
    if example is not None:
        sheet_measurements(wb, example, title="Example")
    sheet_waveforms(wb)

    out = ROOT / "docs" / "model_input_template.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"wrote {out.relative_to(ROOT)}")
    print(f"  sheets: {', '.join(wb.sheetnames)}")
    print(f"  {len(F.REQUIRED_INPUT)} measurement columns")


if __name__ == "__main__":
    main()
